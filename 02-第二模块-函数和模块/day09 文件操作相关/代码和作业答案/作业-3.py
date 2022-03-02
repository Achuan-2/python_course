"""
3. 读取ini文件内容，按照规则写入到Excel中。

- ini文件内容如下：

    ```ini
    [mysqld]
    datadir=/var/lib/mysql
    socket=/var/lib/mysql/mysql.sock
    log-bin=py-mysql-bin
    character-set-server=utf8
    collation-server=utf8_general_ci
    log-error=/var/log/mysqld.log
    # Disabling symbolic-links is recommended to prevent assorted security risks
    symbolic-links=0
    
    [mysqld_safe]
    log-error=/var/log/mariadb/mariadb.log
    pid-file=/var/run/mariadb/mariadb.pid
    
    [client]
    default-character-set=utf8
    ```

- 读取ini格式的文件，并创建一个excel文件，且为每个节点创建一个sheet，然后将节点下的键值写入到excel中，按照如下格式。
    <img src="assets/image-20201218204922898.png" alt="image-20201218204922898" style="zoom: 33%;" />

    - 首行，字体白色 & 单元格背景色蓝色。
    - 内容均居中。
    - 边框。	
"""


import os
import configparser

import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl import workbook

# 文件路径处理
base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir, 'files', 'my.ini')
target_excel_file_path = os.path.join(base_dir, 'my.xlsx')

# 创建excel且默认创建一个sheet（名称为sheet）
wb = workbook.Workbook()
del wb['Sheet']

# 解析 ini 文件
config = configparser.ConfigParser()
config.read(file_path)


for section in config.sections():
    # 在excel中创建一个sheet，名称为ini文件的节点名称
    sheet = wb.create_sheet(section)

    # 边框和居中（表头和内容都需要）
    side = Side(style="thin", color="000000")
    border = Border(top=side, bottom=side, left=side, right=side)

    align = Alignment(horizontal='center', vertical='center')

    # 为此在sheet设置表头
    title_dict = {"A1": "键", "B1": "值"}
    for position, text in title_dict.items():
        cell = sheet[position]
        # 设置值
        cell.value = text
        # 设置居中
        cell.alignment = align
        # 设置背景色
        cell.fill = PatternFill("solid", fgColor="6495ED")
        # 设置字体颜色
        cell.font = Font(name="微软雅黑", color="FFFFFF")
        # 设置边框
        cell.border = border

    # 读取此节点下的所有键值，并将键值写入到当前sheet中
    # row_index = 2
    # for key, val in config.items(section):
    #     c1 = sheet.cell(row_index, 1)
    #     c1.value = key
    #     c1.alignment = align
    #     c1.border = border
    #
    #     c2 = sheet.cell(row_index, 2)
    #     c2.value = val
    #     c2.alignment = align
    #     c2.border = border
    #     row_index += 1

    row_index = 2
    for group in config.items(section):
        # group = ("datadir","/var/lib/mysql")
        for col, text in enumerate(group, 1):
            cell = sheet.cell(row_index, col)
            cell.alignment = align
            cell.border = border
            cell.value = text
        row_index += 1

wb.save(target_excel_file_path)
