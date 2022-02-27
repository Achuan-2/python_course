import os
import hashlib
from datetime import datetime
import random
import string

from openpyxl import load_workbook, workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = "db.xlsx"

def run():
    while True:
        username = input("请输入用户名：")
        if username.upper() == "Q":
            break
        password = input("请输入密码：")
        register(username, password)


def sha256(origin,salt_length=8):
    salt=''.join(random.sample(string.ascii_letters + string.digits, salt_length))
    hash_object = hashlib.sha256(salt.encode('utf-8'))
    hash_object.update(origin.encode('utf-8'))
    return hash_object.hexdigest(),salt


def register(username, password):
    db_file_path = os.path.join(BASE_DIR, FILE_NAME)
    if os.path.exists(db_file_path):
        wb = load_workbook(db_file_path)
        sheet = wb.worksheets[0]
        next_row_position = sheet.max_row + 1
    else:
        wb = workbook.Workbook()
        sheet = wb.worksheets[0]
        next_row_position = 1

    user = sheet.cell(next_row_position, 1)
    user.value = username

    pwd = sheet.cell(next_row_position, 2)
    pwd.value, salt = sha256(password)

    ctime = sheet.cell(next_row_position, 3)
    ctime.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    hash_salt = sheet.cell(next_row_position, 4)
    hash_salt.value = salt
    wb.save(db_file_path)




if __name__ == '__main__':
    run()
