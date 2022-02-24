"""
2. 补充代码：实现去网上获取指定地区的天气信息，并写入到Excel中。
"""

import os 
import requests
from xml.etree import ElementTree as ET
from openpyxl import workbook

# 处理文件路径
base_dir = os.path.dirname(os.path.abspath(__file__))
target_excel_file_path = os.path.join(base_dir, 'weather.xlsx')

# 创建excel且默认会创建一个sheet（名称为Sheet）
wb = workbook.Workbook()
del wb['Sheet']

while True:
    city = input("请输入城市（Q/q退出）：")
    if city.upper()  == "Q":
        break
    url = "http://ws.webxml.com.cn//WebServices/WeatherWebService.asmx/getWeatherbyCityName?theCityName={}".format(city)
    res = requests.get(url=url)
    # 1.提取XML格式中的数据
    root = ET.XML(res.text)

    # 2.为每个城市创建一个sheet，并将获取的xml格式中的数据写入到excel中。
    sheet = wb.create_sheet(city)

    for row_index, node in enumerate(root, 1):
        cell = sheet.cell(row_index, 1)
        cell.value = node.text
wb.save(target_excel_file_path)
